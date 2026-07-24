"""Batch Exact Brand service — parsing, validation, confirmation, and execution.

Reads `docs/batch-exact-brand-production-plan-v1.md` for authoritative rules.
"""

import csv
import hashlib
import io
import re
import unicodedata
from uuid import UUID, uuid4

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.modules.models import (
    AuditLog,
    BatchImport,
    ExactBrandTarget,
    SearchTask,
    SystemSetting,
    TaskVendorPlan,
)
from app.pipeline.outbox import add_event as emit
from app.pipeline.state_machine import transition_task
from app.shared.enums import TaskStatus
from app.shared.models import utc_now

# ── Constants ────────────────────────────────────────────────────────────────

TEMPLATE_VERSION = "exact-brand-import-v1"
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_ROWS = 5000

REQUIRED_COLUMNS = ["company_name", "official_domain"]
OPTIONAL_COLUMNS = ["country", "external_id", "notes"]
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

# Ordered column names for the downloadable template
TEMPLATE_COLUMNS = ["company_name", "official_domain", "country", "external_id", "notes"]

# Valid execution statuses
EXECUTION_PENDING = "pending"
EXECUTION_QUEUED = "queued"
EXECUTION_RUNNING = "running"
EXECUTION_COMPLETED = "completed"
EXECUTION_NO_MATCH = "no_match"
EXECUTION_PARTIAL = "partial"
EXECUTION_RETRYABLE = "retryable"
EXECUTION_FAILED = "failed"
EXECUTION_CANCELLED = "cancelled"

EXECUTION_TERMINAL = {
    EXECUTION_COMPLETED,
    EXECUTION_NO_MATCH,
    EXECUTION_PARTIAL,
    EXECUTION_FAILED,
    EXECUTION_CANCELLED,
}

# BatchImport statuses
BATCH_UPLOADED = "uploaded"
BATCH_PARSING = "parsing"
BATCH_READY = "ready"
BATCH_CONFIRMED = "confirmed"
BATCH_EXECUTING = "executing"
BATCH_COMPLETED = "completed"
BATCH_PARTIAL = "partial"
BATCH_FAILED = "failed"
BATCH_CANCELLED = "cancelled"
BATCH_INVALID = "invalid"

# ── Normalization helpers ────────────────────────────────────────────────────


def _nfkc(text: str) -> str:
    return unicodedata.normalize("NFKC", text)


def _normalize_company_name(raw: str) -> str:
    """NFKC + strip + collapse whitespace."""
    return re.sub(r"\s+", " ", _nfkc(raw).strip())


def _extract_domain(raw: str) -> str:
    """Extract domain from raw input: strip protocol, www, path, query, fragment.

    Returns lowercase IDNA-normalized domain or empty string.
    """
    value = _nfkc(raw).strip().rstrip(".")
    if not value:
        return ""
    # If it looks like an email, reject
    if "@" in value:
        return ""
    if "://" in value and value.split("://", 1)[0].lower() not in {"http", "https"}:
        return ""
    # Add scheme if missing for urlparse
    if "://" not in value:
        value = "https://" + value
    try:
        from urllib.parse import urlparse

        parsed = urlparse(value)
    except Exception:
        return ""
    domain = (parsed.hostname or parsed.netloc or "").lower().strip()
    # Remove www. prefix
    domain = re.sub(r"^www\.", "", domain)
    # Remove any remaining path/port
    domain = domain.split(":")[0]
    domain = domain.split("/")[0]
    domain = domain.strip(".")
    try:
        return domain.encode("idna").decode("ascii")
    except UnicodeError:
        return ""


def _is_valid_domain(domain: str) -> bool:
    """Reject empty, IP addresses, localhost, and invalid TLDs."""
    if not domain:
        return False
    # Reject IP addresses
    ip_pattern = r"^(\d{1,3}\.){3}\d{1,3}$"
    if re.match(ip_pattern, domain):
        return False
    # Reject localhost and variants
    if domain in ("localhost", "127.0.0.1", "0.0.0.0", "::1"):
        return False
    # Must have at least one dot (TLD)
    if "." not in domain:
        return False
    # Reject domains with invalid characters
    if not re.match(r"^[a-z0-9]([a-z0-9-]*[a-z0-9])?(\.[a-z0-9]([a-z0-9-]*[a-z0-9])?)+$", domain):
        return False
    return True


def _is_formula_injection(value: str) -> bool:
    """Detect CSV formula injection prefixes."""
    dangerous = ("=", "+", "-", "@", "\t", "\r", "\n")
    stripped = value.strip()
    for prefix in dangerous:
        if stripped.startswith(prefix):
            return True
    return False


def _csv_safe(value: object) -> str:
    text = str(value or "")
    return f"'{text}" if _is_formula_injection(text) else text


# ── File parsing ─────────────────────────────────────────────────────────────


def parse_csv(file_content: bytes, filename: str) -> dict:
    """Parse CSV/UTF-8 bytes into raw rows. Returns {rows, errors}."""
    # Try UTF-8 first, then UTF-8 BOM
    try:
        text = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = file_content.decode("utf-8")
        except UnicodeDecodeError:
            return {"rows": [], "errors": [{"code": "ENCODING_ERROR", "message": "文件编码不支持，请使用 UTF-8 编码的 CSV 文件"}]}

    lines = text.splitlines()
    if len(lines) < 2:
        return {"rows": [], "errors": [{"code": "EMPTY_FILE", "message": "CSV 文件为空或仅包含表头"}]}

    try:
        reader = csv.DictReader(lines)
    except Exception:
        # Fallback: try with different delimiter detection
        try:
            reader = csv.DictReader(lines)
        except Exception:
            return {"rows": [], "errors": [{"code": "PARSE_ERROR", "message": "无法解析 CSV 文件"}]}

    if reader.fieldnames is None:
        return {"rows": [], "errors": [{"code": "NO_HEADERS", "message": "CSV 文件缺少表头"}]}

    # Normalize header names
    fieldnames = [_nfkc(h).strip().lower().replace(" ", "_") for h in reader.fieldnames]
    missing = [c for c in REQUIRED_COLUMNS if c not in fieldnames]
    if missing:
        return {
            "rows": [],
            "errors": [
                {
                    "code": "MISSING_COLUMNS",
                    "message": f"缺少必填列: {', '.join(missing)}。模板应包含: {', '.join(TEMPLATE_COLUMNS)}",
                }
            ],
        }

    # Map column indices
    col_map = {col: fieldnames.index(col) for col in REQUIRED_COLUMNS}
    for col in OPTIONAL_COLUMNS:
        if col in fieldnames:
            col_map[col] = fieldnames.index(col)

    raw_rows: list[dict] = []
    errors: list[dict] = []
    seen_domains: dict[str, int] = {}

    for row_idx, row in enumerate(reader):
        row_num = row_idx + 2  # 1-indexed, row 1 is header
        if row_num > MAX_ROWS:
            errors.append({"code": "ROW_LIMIT_EXCEEDED", "message": f"超过最大行数限制 ({MAX_ROWS})，已截断"})
            break

        company_name = (row.get("company_name") or row.get("Company Name") or "").strip()
        official_domain = (row.get("official_domain") or row.get("Official Domain") or "").strip()

        # Skip completely empty rows
        if not company_name and not official_domain:
            continue

        row_parse_errors: list[str] = []
        # Formula injection check
        if _is_formula_injection(company_name):
            errors.append({"row": row_num, "code": "FORMULA_INJECTION", "field": "company_name", "message": "公司名称包含潜在公式注入字符"})
            company_name = ""
            row_parse_errors.append("FORMULA_INJECTION")
        if _is_formula_injection(official_domain):
            errors.append({"row": row_num, "code": "FORMULA_INJECTION", "field": "official_domain", "message": "域名包含潜在公式注入字符"})
            official_domain = ""
            row_parse_errors.append("FORMULA_INJECTION")

        country = (row.get("country") or row.get("Country") or "").strip() or None
        external_id = (row.get("external_id") or row.get("External ID") or "").strip() or None
        notes = (row.get("notes") or row.get("Notes") or "").strip() or None

        raw_rows.append({
            "row_number": row_num,
            "company_name_raw": row.get("company_name", ""),
            "official_domain_raw": row.get("official_domain", ""),
            "company_name": company_name,
            "official_domain": official_domain,
            "country": country,
            "external_id": external_id,
            "notes": notes,
            "parse_errors": sorted(set(row_parse_errors)),
        })

        # Track duplicate domains
        if official_domain:
            norm_domain = _extract_domain(official_domain)
            if norm_domain:
                if norm_domain in seen_domains:
                    seen_domains[norm_domain] += 1
                else:
                    seen_domains[norm_domain] = 1

    return {
        "rows": raw_rows,
        "errors": errors,
        "domain_frequencies": seen_domains,
    }


def parse_xlsx(file_content: bytes, filename: str) -> dict:
    """Parse XLSX bytes using openpyxl. Returns same shape as parse_csv."""
    try:
        import openpyxl
    except ImportError:
        return {"rows": [], "errors": [{"code": "XLSX_NOT_SUPPORTED", "message": "服务端未安装 openpyxl，无法解析 .xlsx 文件"}]}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=False, data_only=False)
    except Exception:
        return {"rows": [], "errors": [{"code": "XLSX_PARSE_ERROR", "message": "无法解析 .xlsx 文件，请确认文件格式正确"}]}

    ws = wb.active
    if ws is None:
        return {"rows": [], "errors": [{"code": "EMPTY_WORKSHEET", "message": "Excel 文件中没有找到工作表"}]}

    if ws.merged_cells.ranges:
        wb.close()
        return {
            "rows": [],
            "errors": [{"code": "MERGED_CELLS", "message": "Excel templates cannot contain merged cells"}],
        }

    # Read header row
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [(_nfkc(str(v)).strip().lower().replace(" ", "_") if v else "") for v in row]
        break

    if not headers:
        return {"rows": [], "errors": [{"code": "NO_HEADERS", "message": "Excel 文件缺少表头"}]}

    for index, header in enumerate(headers, start=1):
        if header and ws.column_dimensions[openpyxl.utils.get_column_letter(index)].hidden:
            wb.close()
            return {
                "rows": [],
                "errors": [{"code": "HIDDEN_COLUMN", "message": "Excel business columns cannot be hidden"}],
            }

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        return {
            "rows": [],
            "errors": [
                {
                    "code": "MISSING_COLUMNS",
                    "message": f"缺少必填列: {', '.join(missing)}。模板应包含: {', '.join(TEMPLATE_COLUMNS)}",
                }
            ],
        }

    # Map columns
    col_map: dict[str, int] = {}
    for col in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
        try:
            col_map[col] = headers.index(col)
        except ValueError:
            col_map[col] = -1

    raw_rows: list[dict] = []
    errors: list[dict] = []
    seen_domains: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx - 1 > MAX_ROWS:
            errors.append({"code": "ROW_LIMIT_EXCEEDED", "message": f"超过最大行数限制 ({MAX_ROWS})，已截断"})
            break

        values = list(row) if row else []

        def _cell_value(col_name: str) -> str:
            idx = col_map.get(col_name, -1)
            if 0 <= idx < len(values) and values[idx] is not None:
                return str(values[idx]).strip()
            return ""

        company_name = _cell_value("company_name")
        official_domain = _cell_value("official_domain")
        if not company_name and not official_domain:
            continue

        company_cell = ws.cell(row=row_idx, column=col_map["company_name"] + 1)
        domain_cell = ws.cell(row=row_idx, column=col_map["official_domain"] + 1)
        row_parse_errors: list[str] = []
        if company_cell.data_type == "f" or domain_cell.data_type == "f":
            errors.append({
                "row": row_idx,
                "code": "FORMULA_INJECTION",
                "message": "Excel business fields cannot contain formulas",
            })
            company_name = ""
            official_domain = ""
            row_parse_errors.append("FORMULA_INJECTION")

        if _is_formula_injection(company_name):
            errors.append({"row": row_idx, "code": "FORMULA_INJECTION", "field": "company_name", "message": "公司名称包含潜在公式注入字符"})
            company_name = ""
            row_parse_errors.append("FORMULA_INJECTION")
        if _is_formula_injection(official_domain):
            errors.append({"row": row_idx, "code": "FORMULA_INJECTION", "field": "official_domain", "message": "域名包含潜在公式注入字符"})
            official_domain = ""
            row_parse_errors.append("FORMULA_INJECTION")

        raw_rows.append({
            "row_number": row_idx,
            "company_name_raw": company_name,
            "official_domain_raw": official_domain,
            "company_name": company_name,
            "official_domain": official_domain,
            "country": _cell_value("country") or None,
            "external_id": _cell_value("external_id") or None,
            "notes": _cell_value("notes") or None,
            "parse_errors": sorted(set(row_parse_errors)),
        })

        if official_domain:
            norm_domain = _extract_domain(official_domain)
            if norm_domain:
                seen_domains[norm_domain] = seen_domains.get(norm_domain, 0) + 1

    wb.close()
    return {"rows": raw_rows, "errors": errors, "domain_frequencies": seen_domains}


# ── Validation and normalization ─────────────────────────────────────────────


def _validate_and_normalize_rows(raw_rows: list[dict]) -> dict:
    """Validate and normalize raw parsed rows. Returns validated rows + summary."""
    valid_rows: list[dict] = []
    all_errors: list[dict] = []
    seen_domains: dict[str, int] = {}
    seen_domain_names: dict[str, str] = {}  # domain -> company_name for conflict detection

    for row in raw_rows:
        row_num = row["row_number"]
        row_errors: list[dict] = []
        warnings: list[dict] = []

        for code in row.get("parse_errors", []):
            row_errors.append({"code": code, "message": "Business fields contain a formula"})

        # Normalize company name
        company_name = row.get("company_name", "")
        if not company_name:
            row_errors.append({"code": "MISSING_COMPANY_NAME", "message": "公司名称为空"})

        normalized_name = _normalize_company_name(company_name)
        if normalized_name and len(normalized_name) < 2:
            row_errors.append({"code": "COMPANY_NAME_TOO_SHORT", "message": "公司名称过短"})
        if normalized_name and len(normalized_name) > 500:
            warnings.append({"code": "COMPANY_NAME_TOO_LONG", "message": "公司名称过长，已截断至 500 字符"})
            normalized_name = normalized_name[:500]

        # Normalize domain
        official_domain = row.get("official_domain", "")
        if not official_domain:
            row_errors.append({"code": "MISSING_DOMAIN", "message": "域名为空"})

        normalized_domain = _extract_domain(official_domain)
        if official_domain and not normalized_domain:
            row_errors.append({"code": "INVALID_DOMAIN", "message": f"无法从 \"{official_domain}\" 提取有效域名"})
        elif normalized_domain and not _is_valid_domain(normalized_domain):
            row_errors.append({"code": "INVALID_DOMAIN", "message": f"域名 \"{normalized_domain}\" 无效（不支持 IP 地址或本地地址）"})
        elif normalized_domain:
            # Check for duplicates
            if normalized_domain in seen_domains:
                existing_name = seen_domain_names.get(normalized_domain, "")
                if existing_name and existing_name != normalized_name:
                    row_errors.append({
                        "code": "DUPLICATE_DOMAIN",
                        "message": f"域名 \"{normalized_domain}\" 与第 {seen_domains[normalized_domain]} 行重复",
                    })
                else:
                    row_errors.append({
                        "code": "DUPLICATE_DOMAIN",
                        "message": f"域名 \"{normalized_domain}\" 重复",
                    })
            else:
                seen_domains[normalized_domain] = row_num
                seen_domain_names[normalized_domain] = normalized_name

        # Determine validation status
        if row_errors:
            validation_status = "error"
        elif warnings:
            validation_status = "warning"
        else:
            validation_status = "valid"

        validated_row = {
            "row_number": row_num,
            "company_name": company_name,
            "normalized_company_name": normalized_name,
            "official_domain": official_domain,
            "normalized_domain": normalized_domain,
            "country": row.get("country"),
            "external_id": row.get("external_id"),
            "notes": row.get("notes"),
            "validation_status": validation_status,
            "validation_errors": [e["code"] for e in row_errors],
            "error_details": row_errors,
            "warnings": [w["code"] for w in warnings],
        }
        valid_rows.append(validated_row)
        all_errors.extend(row_errors)

    # Error summary by code
    error_summary: dict[str, int] = {}
    for e in all_errors:
        error_summary[e["code"]] = error_summary.get(e["code"], 0) + 1

    total = len(valid_rows)
    error_rows = sum(1 for r in valid_rows if r["validation_status"] == "error")
    warning_rows = sum(1 for r in valid_rows if r["validation_status"] == "warning")
    valid_only = total - error_rows - warning_rows
    duplicate_rows = sum(1 for r in valid_rows if "DUPLICATE_DOMAIN" in r.get("validation_errors", []))

    return {
        "rows": valid_rows,
        "total_rows": total,
        "valid_rows": valid_only,
        "warning_rows": warning_rows,
        "invalid_rows": error_rows,
        "duplicate_rows": duplicate_rows,
        "error_summary": error_summary,
    }


# ── Preview (no vendor calls, no persistence) ────────────────────────────────


def build_preview(filename: str, file_content: bytes) -> dict:
    """Parse, validate, and return a read-only preview. Makes zero Vendor calls."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        parse_result = parse_csv(file_content, filename)
    elif ext == "xlsx":
        parse_result = parse_xlsx(file_content, filename)
    else:
        return {"error": {"code": "UNSUPPORTED_FORMAT", "message": "仅支持 .xlsx 和 .csv 文件格式"}}

    if parse_result.get("errors") and not parse_result.get("rows"):
        return {"error": parse_result["errors"][0]}

    raw_rows = parse_result["rows"]
    validation = _validate_and_normalize_rows(raw_rows)

    file_hash = hashlib.sha256(file_content).hexdigest()

    return {
        "filename": filename,
        "template_version": TEMPLATE_VERSION,
        "file_hash": file_hash,
        "total_rows": validation["total_rows"],
        "valid_rows": validation["valid_rows"],
        "warning_rows": validation["warning_rows"],
        "invalid_rows": validation["invalid_rows"],
        "duplicate_rows": validation["duplicate_rows"],
        "rows": validation["rows"],
        "error_summary": validation["error_summary"],
        "limits": {
            "max_rows": MAX_ROWS,
            "max_file_size_mb": MAX_FILE_SIZE_BYTES // (1024 * 1024),
        },
    }


# ── Template download ────────────────────────────────────────────────────────


def generate_template_csv() -> bytes:
    """Generate the downloadable CSV template."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(TEMPLATE_COLUMNS)
    # Add one example row
    writer.writerow(["MANGO", "mango.com", "Spain", "C-0001", "Key account"])
    return output.getvalue().encode("utf-8-sig")


# ── BatchImport persistence ──────────────────────────────────────────────────


def create_batch_import(
    db: Session,
    *,
    filename: str,
    file_hash: str,
    parsed_rows: list[dict],
    organization_id: UUID | None = None,
    organization_unit_id: UUID | None = None,
    created_by: UUID | None = None,
) -> BatchImport:
    """Persist a BatchImport from preview data. Does NOT create targets or call vendors."""
    # Check if this file_hash already has a batch for this org
    organization_clause = (
        BatchImport.organization_id == organization_id
        if organization_id
        else BatchImport.organization_id.is_(None)
    )
    existing = db.scalar(
        select(BatchImport).where(BatchImport.file_hash == file_hash, organization_clause)
    )
    if existing is not None:
        return existing

    # Classify rows
    valid_rows_list = [r for r in parsed_rows if r["validation_status"] in ("valid", "warning")]
    invalid_count = sum(1 for r in parsed_rows if r["validation_status"] == "error")
    duplicate_count = sum(1 for r in parsed_rows if "DUPLICATE_DOMAIN" in r.get("validation_errors", []))
    warning_count = sum(1 for r in parsed_rows if r["validation_status"] == "warning")

    error_summary: dict[str, int] = {}
    for r in parsed_rows:
        for e in r.get("validation_errors", []):
            error_summary[e] = error_summary.get(e, 0) + 1

    batch = BatchImport(
        id=uuid4(),
        organization_id=organization_id,
        department_id=organization_unit_id,
        owner_id=created_by,
        created_by=created_by,
        filename=filename,
        template_version=TEMPLATE_VERSION,
        file_hash=file_hash,
        status=BATCH_READY,
        total_rows=len(parsed_rows),
        valid_rows=len(valid_rows_list),
        warning_rows=warning_count,
        invalid_rows=invalid_count,
        duplicate_rows=duplicate_count,
        error_summary=error_summary,
        parsed_preview={"rows": parsed_rows},
    )
    db.add(batch)
    db.flush()
    return batch


# ── Confirmation (idempotent) ────────────────────────────────────────────────


def confirm_batch_import(
    db: Session,
    *,
    batch_id: UUID,
    config,  # BatchImportConfirm
    organization_id: UUID | None = None,
    user_id: UUID | None = None,
) -> dict:
    """Idempotent confirm: creates parent SearchTask, Targets, TaskVendorPlan, and Outbox events.

    Returns the batch with parent_task_id set. Repeated calls return the same task.
    """
    batch = db.get(BatchImport, batch_id)
    if batch is None:
        raise ValueError(f"BatchImport {batch_id} not found")
    if batch.status not in (BATCH_READY, BATCH_CONFIRMED, BATCH_EXECUTING):
        raise ValueError(f"BatchImport {batch_id} is in status {batch.status}, cannot confirm")

    # Idempotent: if already confirmed, return existing parent task
    if batch.parent_task_id is not None:
        parent_task = db.get(SearchTask, batch.parent_task_id)
        if parent_task is not None:
            targets = list(
                db.scalars(
                    select(ExactBrandTarget).where(
                        ExactBrandTarget.batch_import_id == batch.id
                    )
                ).all()
            )
            return {
                "batch": batch,
                "parent_task": parent_task,
                "targets": targets,
                "already_confirmed": True,
            }

    parsed_preview = batch.parsed_preview or {}
    parsed_rows = parsed_preview.get("rows", [])
    valid_rows = [r for r in parsed_rows if r.get("validation_status") in ("valid", "warning")]

    if not valid_rows:
        raise ValueError("No valid rows to confirm")

    # Create parent SearchTask
    parent_task_id = uuid4()
    now = utc_now()

    # Capture configuration snapshot (secret-free)
    config_snapshot = {
        "mode": "batch_exact_brand",
        "selected_vendors": config.selected_vendors,
        "target_titles": config.target_titles,
        "contacts_limit_per_brand": config.contacts_limit_per_brand,
        "reliable_email_only": config.reliable_email_only,
        "skip_existing_brands": config.skip_existing_brands,
        "budget_limit": config.budget_limit,
        "max_concurrency": config.max_concurrency,
        "retry_limit_per_target": config.retry_limit_per_target,
        "batch_import_id": str(batch.id),
        "template_version": batch.template_version,
        "total_targets": len(valid_rows),
        "created_at": now.isoformat(),
    }

    vendor_plan_id = uuid4()

    parent_task = SearchTask(
        id=parent_task_id,
        name=config.name,
        mode="batch_exact_brand",
        status=TaskStatus.draft,
        filters={
            "batch_import_id": str(batch.id),
            "selected_vendors": config.selected_vendors,
            "target_titles": config.target_titles,
            "contacts_limit_per_brand": config.contacts_limit_per_brand,
        },
        configuration_version=TEMPLATE_VERSION,
        configuration_snapshot=config_snapshot,
        pipeline_version="1.0.0",
        organization_id=organization_id,
        budget_limit=config.budget_limit,
        trace_id=str(uuid4()),
    )
    db.add(parent_task)
    db.flush()

    # Create TaskVendorPlan
    primary_vendor = config.selected_vendors[0] if config.selected_vendors else "apollo"
    execution_mode = (
        "apollo_hunter" if len(config.selected_vendors) == 2
        else "apollo_only" if config.selected_vendors[0] == "apollo"
        else "hunter_only"
    )

    vendor_plan = TaskVendorPlan(
        id=vendor_plan_id,
        task_id=parent_task_id,
        primary_vendor=primary_vendor,
        fallback_vendors=[],
        verification_vendor=None,
        adapter_version="v1",
        execution_mode=execution_mode,
        selected_vendors=config.selected_vendors,
        pipeline_source="user_selection",
        vendor_routes={
            v: {
                "company_search": v,
                "contact_search": v,
                "email_finder": v,
                "email_verifier": v,
            }
            for v in config.selected_vendors
        },
    )
    db.add(vendor_plan)

    # Create ExactBrandTargets for each valid row
    targets = []
    for row in valid_rows:
        target = ExactBrandTarget(
            id=uuid4(),
            batch_import_id=batch.id,
            search_task_id=parent_task_id,
            organization_id=organization_id,
            row_number=row["row_number"],
            external_id=row.get("external_id"),
            company_name=row.get("company_name", ""),
            normalized_company_name=row.get("normalized_company_name", ""),
            official_domain=row.get("official_domain", ""),
            normalized_domain=row.get("normalized_domain", ""),
            country=row.get("country"),
            notes=row.get("notes"),
            raw_input={
                "company_name": row.get("company_name", ""),
                "official_domain": row.get("official_domain", ""),
                "country": row.get("country"),
                "external_id": row.get("external_id"),
                "notes": row.get("notes"),
            },
            validation_status=row.get("validation_status", "valid"),
            validation_errors=row.get("validation_errors", []),
            execution_status=EXECUTION_PENDING,
            max_attempts=config.retry_limit_per_target,
        )
        db.add(target)
        targets.append(target)

    # Transition parent task
    transition_task(parent_task, TaskStatus.queued)
    parent_task.queued_at = now

    # Update batch
    batch.status = BATCH_CONFIRMED
    batch.parent_task_id = parent_task_id
    batch.confirmed_at = now

    # Emit domain events
    emit(db, "batch_exact_brand_confirmed", "batch_import", str(batch.id), {
        "batch_id": str(batch.id),
        "parent_task_id": str(parent_task_id),
        "organization_id": str(organization_id) if organization_id else None,
        "target_count": len(targets),
        "vendors": config.selected_vendors,
        "actor_id": str(user_id) if user_id else None,
    })

    emit(db, "search_task_created", "search_task", str(parent_task_id), {
        "task_id": str(parent_task_id),
        "mode": "batch_exact_brand",
        "target_count": len(targets),
        "batch_id": str(batch.id),
    })

    # Audit log
    audit = AuditLog(
        id=uuid4(),
        actor_id=str(user_id) if user_id else None,
        action="batch_exact_brand.confirm",
        entity_type="batch_import",
        entity_id=str(batch.id),
        before={"status": BATCH_READY},
        after={
            "status": BATCH_CONFIRMED,
            "parent_task_id": str(parent_task_id),
            "target_count": len(targets),
        },
    )
    db.add(audit)

    db.flush()

    return {
        "batch": batch,
        "parent_task": parent_task,
        "targets": targets,
        "already_confirmed": False,
    }


# ── Target status aggregation ────────────────────────────────────────────────


def get_batch_target_summary(db: Session, batch_id: UUID) -> dict:
    """Return aggregated counts for a batch's targets."""
    targets = list(
        db.scalars(
            select(ExactBrandTarget).where(ExactBrandTarget.batch_import_id == batch_id)
        ).all()
    )

    summary = {
        "targets_total": len(targets),
        "targets_completed": 0,
        "targets_running": 0,
        "targets_pending": 0,
        "targets_no_match": 0,
        "targets_failed": 0,
        "targets_cancelled": 0,
        "total_reliable_emails": 0,
        "total_review_emails": 0,
    }

    for t in targets:
        if t.execution_status == EXECUTION_COMPLETED:
            summary["targets_completed"] += 1
        elif t.execution_status == EXECUTION_RUNNING:
            summary["targets_running"] += 1
        elif t.execution_status == EXECUTION_PENDING or t.execution_status == EXECUTION_QUEUED:
            summary["targets_pending"] += 1
        elif t.execution_status == EXECUTION_NO_MATCH:
            summary["targets_no_match"] += 1
        elif t.execution_status == EXECUTION_FAILED or t.execution_status == EXECUTION_RETRYABLE:
            summary["targets_failed"] += 1
        elif t.execution_status == EXECUTION_CANCELLED:
            summary["targets_cancelled"] += 1
        summary["total_reliable_emails"] += t.reliable_email_count or 0
        summary["total_review_emails"] += t.review_email_count or 0

    return summary


def get_batch_detail(db: Session, batch_id: UUID) -> dict:
    """Get batch detail with aggregated target stats."""
    batch = db.get(BatchImport, batch_id)
    if batch is None:
        raise ValueError(f"BatchImport {batch_id} not found")

    summary = get_batch_target_summary(db, batch_id)

    return {
        "id": batch.id,
        "filename": batch.filename,
        "template_version": batch.template_version,
        "status": batch.status,
        "total_rows": batch.total_rows,
        "valid_rows": batch.valid_rows,
        "warning_rows": batch.warning_rows,
        "invalid_rows": batch.invalid_rows,
        "duplicate_rows": batch.duplicate_rows,
        "parent_task_id": batch.parent_task_id,
        "created_at": batch.created_at.isoformat() if batch.created_at else None,
        "confirmed_at": batch.confirmed_at.isoformat() if batch.confirmed_at else None,
        **summary,
    }


def get_targets_for_task(
    db: Session,
    task_id: UUID,
    *,
    page: int = 1,
    page_size: int = 50,
    status_filter: str | None = None,
) -> dict:
    """Paginated target list for a parent task."""
    base = select(ExactBrandTarget).where(ExactBrandTarget.search_task_id == task_id)
    if status_filter:
        base = base.where(ExactBrandTarget.execution_status == status_filter)

    total = db.scalar(select(func.count()).select_from(base.subquery()))
    rows = list(
        db.scalars(base.order_by(ExactBrandTarget.row_number).offset((page - 1) * page_size).limit(page_size)).all()
    )

    all_targets = list(
        db.scalars(
            select(ExactBrandTarget).where(ExactBrandTarget.search_task_id == task_id)
        ).all()
    )
    summary = {
        "total": len(all_targets),
        "completed": sum(t.execution_status == EXECUTION_COMPLETED for t in all_targets),
        "running": sum(t.execution_status == EXECUTION_RUNNING for t in all_targets),
        "pending": sum(
            t.execution_status in (EXECUTION_PENDING, EXECUTION_QUEUED) for t in all_targets
        ),
        "no_match": sum(t.execution_status == EXECUTION_NO_MATCH for t in all_targets),
        "failed": sum(
            t.execution_status in (EXECUTION_FAILED, EXECUTION_RETRYABLE) for t in all_targets
        ),
        "partial": sum(t.execution_status == EXECUTION_PARTIAL for t in all_targets),
        "cancelled": sum(t.execution_status == EXECUTION_CANCELLED for t in all_targets),
        "reliable_emails": sum(t.reliable_email_count or 0 for t in all_targets),
        "review_emails": sum(t.review_email_count or 0 for t in all_targets),
    }

    return {
        "items": rows,
        "total": total or 0,
        "page": page,
        "page_size": page_size,
        "summary": summary,
    }


# ── Retry failed targets ─────────────────────────────────────────────────────


def retry_targets(
    db: Session,
    *,
    task_id: UUID,
    target_ids: list[UUID],
    user_id: UUID | None = None,
) -> int:
    """Retry specific failed/retryable targets. Returns count of targets queued."""
    retried = 0
    for tid in target_ids:
        target = db.get(ExactBrandTarget, tid)
        if target is None:
            continue
        if str(target.search_task_id) != str(task_id):
            continue
        if target.execution_status in (EXECUTION_FAILED, EXECUTION_RETRYABLE):
            if int(target.execution_attempts or 0) >= int(target.max_attempts or 3):
                continue
            target.execution_status = EXECUTION_QUEUED
            target.current_stage = None
            target.error_code = None
            target.error_message = None
            target.lease_owner = None
            target.lease_expires_at = None
            retried += 1

    if retried > 0:
        # Transition parent task back to queued if it's in partial/failed
        parent_task = db.get(SearchTask, task_id)
        if parent_task is not None and parent_task.status in (TaskStatus.partial, TaskStatus.failed):
            transition_task(parent_task, TaskStatus.queued)

        # Audit
        audit = AuditLog(
            id=uuid4(),
            actor_id=str(user_id) if user_id else None,
            action="batch_exact_brand.retry_targets",
            entity_type="search_task",
            entity_id=str(task_id),
            after={"retried_target_ids": [str(t) for t in target_ids], "count": retried},
        )
        db.add(audit)

    return retried


# ── Exports ──────────────────────────────────────────────────────────────────


def export_target_errors(db: Session, task_id: UUID) -> str:
    """Export error rows as CSV string."""
    targets = list(
        db.scalars(
            select(ExactBrandTarget)
            .where(ExactBrandTarget.search_task_id == task_id)
            .where(
                ExactBrandTarget.execution_status.in_(
                    [EXECUTION_FAILED, EXECUTION_RETRYABLE, EXECUTION_NO_MATCH]
                )
            )
            .order_by(ExactBrandTarget.row_number)
        ).all()
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["row_number", "company_name", "official_domain", "execution_status", "error_code", "error_message"])
    for t in targets:
        writer.writerow([
            t.row_number,
            _csv_safe(t.company_name),
            _csv_safe(t.official_domain),
            t.execution_status,
            t.error_code or "",
            _csv_safe(t.error_message),
        ])
    return output.getvalue()


def export_reliable_emails(db: Session, task_id: UUID, scope: str = "verified") -> str:
    """Export verified, reviewable, or all emails across completed targets."""
    from app.modules.models import EmailAddress

    targets = list(
        db.scalars(
            select(ExactBrandTarget).where(
                ExactBrandTarget.search_task_id == task_id,
                ExactBrandTarget.execution_status == EXECUTION_COMPLETED,
            )
        ).all()
    )
    brand_ids = [t.brand_id for t in targets if t.brand_id]

    if not brand_ids:
        return "company_name,domain,email,contact_name,title,authenticity_level\n"

    email_query = select(EmailAddress).where(EmailAddress.brand_id.in_(brand_ids))
    if scope == "verified":
        email_query = email_query.where(EmailAddress.authenticity_level == "verified")
    elif scope == "reviewable":
        email_query = email_query.where(
            EmailAddress.authenticity_level.in_(["verified", "probable", "risky", "unverified"])
        )
    emails = list(db.scalars(email_query).all())

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["company_name", "domain", "email", "contact_name", "title", "authenticity_level"])

    # Build lookup
    brand_map: dict[str, ExactBrandTarget] = {str(t.brand_id): t for t in targets if t.brand_id}
    for e in emails:
        target = brand_map.get(str(e.brand_id))
        writer.writerow([
            _csv_safe(target.company_name if target else ""),
            _csv_safe(target.normalized_domain if target else ""),
            _csv_safe(e.address),
            "",  # contact name can be joined if needed
            "",
            e.authenticity_level,
        ])
    return output.getvalue()


# ── Target execution ─────────────────────────────────────────────────────────


def _retired_execute_target_pipeline(
    db: Session,
    target: ExactBrandTarget,
) -> dict:
    """Execute the vendor pipeline for a single ExactBrandTarget.

    Reuses existing service functions — no duplicate adapter logic.
    """
    parent_task = db.get(SearchTask, target.search_task_id)
    if parent_task is None:
        raise ValueError(f"Parent task {target.search_task_id} not found for target {target.id}")

    target.execution_status = EXECUTION_RUNNING
    target.current_stage = "company_search"
    db.flush()

    config = parent_task.configuration_snapshot or {}
    vendors = config.get("selected_vendors", ["apollo"])
    target_titles = config.get("target_titles", [])
    contacts_limit = config.get("contacts_limit_per_brand", 5)

    all_contacts: list[dict] = []
    all_emails: list[dict] = []
    company_found = False
    brand_id: str | None = None
    errors: list[str] = []

    from app.modules.services import (
        _ensure_email_verified,
        _title_matches_targets,
        create_brand,
        create_contact,
        create_email,
        enabled_providers,
        execute_provider_waterfall,
        get_or_create_company,
    )
    from app.modules.schemas import BrandCreate, ContactCreate, EmailCreate

    for vendor in vendors:
        if parent_task.status == TaskStatus.cancelled:
            target.execution_status = EXECUTION_CANCELLED
            return {"cancelled": True}

        try:
            # ── Company search ────────────────────────────────────────────
            company_providers = enabled_providers(db, "company_search", allowed_vendors={vendor})
            if not company_providers:
                errors.append(f"No enabled company_search provider for {vendor}")
                continue

            payload = {
                "organization_name": target.normalized_company_name,
                "domain": target.normalized_domain,
                "mode": "exact_brand",
            }
            if target.country:
                payload["country"] = target.country

            result = execute_provider_waterfall(
                db, "company_search", company_providers, payload,
                company_filter=None,
                task=parent_task,
            )
            if not result.ok or not result.items:
                errors.append(f"{vendor}: company_search returned no results for {target.normalized_company_name} ({target.normalized_domain})")
                continue

            companies = result.items
            company = companies[0]  # Take first match
            company_found = True

            # Create brand
            domain = company.get("domain") or company.get("website") or target.normalized_domain
            brand_create = BrandCreate(
                name=target.normalized_company_name[:255],
                website=domain[:255] if domain else None,
                country=target.country,
            )
            brand = create_brand(
                db, brand_create,
                organization_id=parent_task.organization_id,
                source_type="commercial_api",
                provider=vendor,
            )
            db.flush()
            brand_id = str(brand.id)
            target.brand_id = UUID(brand_id) if brand_id else None

            # Get or create company
            get_or_create_company(
                db, brand, domain=target.normalized_domain,
                country=target.country, provider=vendor,
            )

            # ── Contact search ────────────────────────────────────────────
            target.current_stage = "contact_search"
            contact_providers = enabled_providers(db, "contact_search", allowed_vendors={vendor})
            if contact_providers:
                contact_payload = {
                    "domain": target.normalized_domain,
                    "person_titles": target_titles,
                    "limit": contacts_limit * 3,
                }
                contact_result = execute_provider_waterfall(
                    db, "contact_search", contact_providers, contact_payload,
                    contact_filter=None,
                    task=parent_task,
                )
                if contact_result.ok and contact_result.items:
                    # Filter contacts by title
                    matched_contacts = []
                    for c in contact_result.items:
                        title = c.get("title") or c.get("job_title") or ""
                        if _title_matches_targets(title, target_titles):
                            matched_contacts.append(c)
                            if len(matched_contacts) >= contacts_limit:
                                break

                    for contact_data in matched_contacts[:contacts_limit]:
                        try:
                            contact = create_contact(
                                db,
                                ContactCreate(
                                    brand_id=UUID(brand_id) if brand_id else None,
                                    first_name=(contact_data.get("first_name") or "Unknown")[:120],
                                    last_name=(contact_data.get("last_name") or "")[:120],
                                    title=(contact_data.get("title") or "")[:255],
                                    linkedin_url=contact_data.get("linkedin_url"),
                                ),
                                organization_id=parent_task.organization_id,
                                provider=vendor,
                            )
                            db.flush()

                            all_contacts.append({
                                "contact_id": str(contact.id),
                                "vendor": vendor,
                                "data": contact_data,
                            })

                            # Extract emails from contact data
                            emails_from_contact = contact_data.get("email") or contact_data.get("emails") or []
                            if isinstance(emails_from_contact, str):
                                emails_from_contact = [emails_from_contact]
                            for email_addr in emails_from_contact:
                                if not email_addr or "@" not in str(email_addr):
                                    continue
                                try:
                                    email_obj = create_email(
                                        db,
                                        EmailCreate(
                                            contact_id=contact.id,
                                            brand_id=UUID(brand_id) if brand_id else None,
                                            address=str(email_addr),
                                        ),
                                        organization_id=parent_task.organization_id,
                                        provider=vendor,
                                    )
                                    all_emails.append({
                                        "email_id": str(email_obj.id),
                                        "address": str(email_addr),
                                        "vendor": vendor,
                                    })
                                except Exception:
                                    pass
                        except Exception as exc:
                            errors.append(f"Contact creation failed: {exc}")

            # ── Email discovery (Hunter email_finder) ─────────────────────
            if vendor == "hunter" and len(all_emails) < contacts_limit:
                target.current_stage = "email_finder"
                email_providers = enabled_providers(db, "email_finder", allowed_vendors={"hunter"})
                if email_providers:
                    for contact_data in all_contacts:
                        try:
                            fn = contact_data.get("data", {}).get("first_name", "")
                            ln = contact_data.get("data", {}).get("last_name", "")
                            if fn and ln:
                                email_result = execute_provider_waterfall(
                                    db, "email_finder", email_providers,
                                    {"first_name": fn, "last_name": ln, "domain": target.normalized_domain},
                                    email_filter=None,
                                    task=parent_task,
                                )
                                if email_result.ok and email_result.items:
                                    for item in email_result.items:
                                        addr = item.get("email") or item.get("address") or ""
                                        if not addr or "@" not in addr:
                                            continue
                                        try:
                                            email_obj = create_email(
                                                db,
                                                EmailCreate(
                                                    contact_id=UUID(contact_data["contact_id"]) if contact_data.get("contact_id") else None,
                                                    brand_id=UUID(brand_id) if brand_id else None,
                                                    address=addr,
                                                ),
                                                organization_id=parent_task.organization_id,
                                                provider=vendor,
                                            )
                                            all_emails.append({
                                                "email_id": str(email_obj.id),
                                                "address": addr,
                                                "vendor": vendor,
                                            })
                                        except Exception:
                                            pass
                        except Exception:
                            pass

        except Exception as exc:
            errors.append(f"{vendor}: {exc}")
            continue

    # ── Apply email verification and update target ────────────────────────
    target.current_stage = "result_persist"
    reliable_count = 0
    review_count = 0

    if all_emails:
        from app.modules.models import EmailAddress as EmailAddr
        for email_info in all_emails:
            try:
                email_obj = db.get(EmailAddr, UUID(email_info["email_id"]) if email_info.get("email_id") else None)
                if email_obj:
                    _ensure_email_verified(db, email_obj, parent_task)
                    if email_obj.authenticity_level == "verified":
                        reliable_count += 1
                    elif email_obj.authenticity_level in ("probable", "risky", "unverified"):
                        review_count += 1
            except Exception:
                pass

    target.contact_count = len(all_contacts)
    target.reliable_email_count = reliable_count
    target.review_email_count = review_count
    target.vendor_results = {
        "vendors_used": vendors,
        "companies_found": 1 if company_found else 0,
        "contacts_found": len(all_contacts),
        "emails_found": len(all_emails),
        "errors": errors[:10],
    }

    if not company_found:
        target.execution_status = EXECUTION_NO_MATCH
        target.current_stage = None
    elif errors and not all_contacts:
        target.execution_status = EXECUTION_FAILED
        target.error_code = "NO_CONTACTS"
        target.error_message = "; ".join(errors[:5])
    else:
        target.execution_status = EXECUTION_COMPLETED
        target.current_stage = None

    return {
        "target_id": str(target.id),
        "status": target.execution_status,
        "brand_id": brand_id,
        "contacts": len(all_contacts),
        "reliable_emails": reliable_count,
        "review_emails": review_count,
        "errors": errors[:10],
    }


def execute_target_pipeline(db: Session, target: ExactBrandTarget) -> dict:
    """Execute one target through the canonical frozen Vendor pipeline."""
    parent_task = db.get(SearchTask, target.search_task_id)
    if parent_task is None:
        raise ValueError(f"Parent task {target.search_task_id} not found for target {target.id}")
    if parent_task.status == TaskStatus.cancelled:
        target.execution_status = EXECUTION_CANCELLED
        target.current_stage = None
        return {"target_id": str(target.id), "status": EXECUTION_CANCELLED}
    if parent_task.status == TaskStatus.paused:
        target.execution_status = EXECUTION_QUEUED
        target.current_stage = None
        return {"target_id": str(target.id), "status": EXECUTION_QUEUED}

    from app.modules.models import Brand, EmailAddress, PipelineStageRun, Website
    from app.pipeline.runner import begin_stage, complete_stage, fail_stage
    from app.pipeline.vendor_pipeline import _persist_vendor_result, execute_vendor_pipeline

    config = parent_task.configuration_snapshot or {}
    if config.get("skip_existing_brands"):
        existing_brand = db.scalar(
            select(Brand)
            .join(Website, Website.brand_id == Brand.id)
            .where(Website.domain == target.normalized_domain, Brand.deleted_at.is_(None))
        )
        if existing_brand is not None:
            target.brand_id = existing_brand.id
            target.execution_status = EXECUTION_COMPLETED
            target.current_stage = None
            target.vendor_results = {"skipped_existing_brand": True}
            return {"target_id": str(target.id), "status": EXECUTION_COMPLETED, "skipped": True}

    spent = float(
        db.scalar(
            select(func.coalesce(func.sum(PipelineStageRun.cost), 0)).where(
                PipelineStageRun.task_id == parent_task.id
            )
        )
        or 0
    )
    if parent_task.budget_limit is not None and spent >= float(parent_task.budget_limit):
        target.execution_status = EXECUTION_RETRYABLE
        target.error_code = "BUDGET_EXHAUSTED"
        target.error_message = "Task budget limit reached before Vendor execution"
        target.current_stage = None
        return {"target_id": str(target.id), "status": EXECUTION_RETRYABLE}

    filters = {
        "mode": "exact_brand",
        "brand_keywords": [target.normalized_company_name],
        "official_domains": [target.normalized_domain],
        "countries": [target.country] if target.country else [],
        "target_titles": list(config.get("target_titles") or []),
        "contacts_limit_per_brand": int(config.get("contacts_limit_per_brand") or 5),
        "brand_limit": 1,
        "selected_vendors": list(config.get("selected_vendors") or ["apollo"]),
    }
    vendors = filters["selected_vendors"]
    persisted: dict[str, set[str]] = {
        "brand_ids": set(), "contact_ids": set(), "email_ids": set()
    }
    vendor_summaries: dict[str, dict] = {}
    errors: list[str] = []

    for vendor in vendors:
        db.refresh(parent_task)
        if parent_task.status in {TaskStatus.cancelled, TaskStatus.paused}:
            target.execution_status = (
                EXECUTION_CANCELLED
                if parent_task.status == TaskStatus.cancelled
                else EXECUTION_QUEUED
            )
            target.current_stage = None
            return {"target_id": str(target.id), "status": target.execution_status}
        target.current_stage = f"{vendor}:provider_search"
        stage_run = begin_stage(
            db,
            parent_task.id,
            "provider_search",
            {
                "target_id": str(target.id),
                "vendor": vendor,
                "mode": "exact_brand",
                "filters": filters,
            },
        )
        try:
            result = execute_vendor_pipeline(db, parent_task, vendor, filters=filters)
            result.stage_run = stage_run
            if result.ok:
                complete_stage(
                    stage_run,
                    {
                        "target_id": str(target.id),
                        "vendor": vendor,
                        "accepted_company_count": len(result.companies),
                        "stage_count": len(result.stages),
                    },
                )
            else:
                fail_stage(
                    stage_run,
                    RuntimeError("; ".join(result.errors) or f"{vendor} pipeline failed"),
                    retryable=True,
                )
            if result.companies:
                summary = _persist_vendor_result(db, parent_task, result)
                for key in persisted:
                    persisted[key].update(summary[key])
            errors.extend(result.errors)
            vendor_summaries[vendor] = {
                "ok": result.ok,
                "company_count": len(result.companies),
                "errors": result.errors[:5],
            }
        except Exception as exc:
            fail_stage(stage_run, exc, retryable=True)
            errors.append(f"{vendor}: {exc}")
            vendor_summaries[vendor] = {"ok": False, "errors": [str(exc)[:500]]}

    email_ids = [UUID(value) for value in persisted["email_ids"]]
    emails = (
        list(db.scalars(select(EmailAddress).where(EmailAddress.id.in_(email_ids))).all())
        if email_ids
        else []
    )
    reliable_count = sum(email.authenticity_level == "verified" for email in emails)
    review_count = sum(
        email.authenticity_level in {"probable", "risky", "unverified"} for email in emails
    )
    brand_ids = sorted(persisted["brand_ids"])
    target.brand_id = UUID(brand_ids[0]) if brand_ids else None
    target.contact_count = len(persisted["contact_ids"])
    target.reliable_email_count = reliable_count
    target.review_email_count = review_count
    target.vendor_results = vendor_summaries
    target.current_stage = None
    successful_vendors = sum(
        bool(item.get("company_count")) for item in vendor_summaries.values()
    )
    if not brand_ids:
        target.execution_status = EXECUTION_NO_MATCH if not errors else EXECUTION_RETRYABLE
    elif errors and successful_vendors < len(vendors):
        target.execution_status = EXECUTION_PARTIAL
    else:
        target.execution_status = EXECUTION_COMPLETED
    target.error_code = (
        "VENDOR_PARTIAL_FAILURE" if target.execution_status == EXECUTION_PARTIAL else None
    )
    target.error_message = "; ".join(errors[:5]) or None
    emit(
        db,
        "batch_exact_brand_target_finished",
        "exact_brand_target",
        str(target.id),
        {
            "task_id": str(parent_task.id),
            "status": target.execution_status,
            "reliable_email_count": reliable_count,
        },
    )
    return {
        "target_id": str(target.id),
        "status": target.execution_status,
        "brand_id": str(target.brand_id) if target.brand_id else None,
        "contacts": target.contact_count,
        "reliable_emails": reliable_count,
        "review_emails": review_count,
        "errors": errors[:10],
    }


def aggregate_parent_task_status(db: Session, task_id: UUID) -> str:
    """Compute the correct parent task status from its targets."""
    targets = list(
        db.scalars(
            select(ExactBrandTarget).where(ExactBrandTarget.search_task_id == task_id)
        ).all()
    )
    if not targets:
        return TaskStatus.completed

    statuses = [t.execution_status for t in targets]
    if all(s == EXECUTION_COMPLETED for s in statuses):
        return TaskStatus.completed
    if all(s == EXECUTION_CANCELLED for s in statuses):
        return TaskStatus.cancelled
    if all(s in (*EXECUTION_TERMINAL, EXECUTION_CANCELLED) for s in statuses):
        if any(s == EXECUTION_COMPLETED for s in statuses):
            return TaskStatus.partial
        if all(s == EXECUTION_FAILED for s in statuses):
            return TaskStatus.failed
        if all(s == EXECUTION_NO_MATCH for s in statuses):
            return TaskStatus.completed  # All "no_match" is a valid result
        return TaskStatus.partial
    if any(s in (EXECUTION_RUNNING, EXECUTION_PENDING, EXECUTION_QUEUED) for s in statuses):
        return TaskStatus.running
    return TaskStatus.partial


# ── Feature flag ─────────────────────────────────────────────────────────────


def is_batch_exact_brand_enabled(db: Session) -> bool:
    """Check if the batch exact brand feature is enabled."""
    setting = db.scalar(
        select(SystemSetting).where(SystemSetting.key == "feature_batch_exact_brand")
    )
    if setting is None:
        return True  # Default enabled
    return setting.value.get("enabled", True) if isinstance(setting.value, dict) else True
